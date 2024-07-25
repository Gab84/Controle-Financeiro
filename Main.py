from customtkinter import *
from PIL import Image, ImageTk

import openpyxl


#cores
fundo = "#4D3D14"
fundo_cima = "#F7EED7"
botoes = "#F4C68F"
fundo_cima_utils= "#A28F61"
#auxiliar



class Janela:
    def __init__(self,username,salario) -> None:
        self.janela = CTk()
        self.username = username
        self.salario = salario
        self.homepage()
          
    def run(self):
        self.janela.mainloop()
    
    def obter_saldo(self):
        
        # Carrega a planilha existente
        wb = openpyxl.load_workbook(f"anotacoes/{self.username}_controle_financeiro.xlsx", data_only=True)
        ws = wb.active
        
        # Lê o valor do salário
        self.salario = ws["I2"].value
        
        # Inicializa as variáveis de receita e despesa
        receita = 0
        despesa = 0
        
        # Itera sobre as linhas da planilha e acumula as receitas e despesas
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=6, max_col=7, values_only=True):
            valor, tipo = row
            if tipo == "Entrada":
                receita += valor
            elif tipo == "Saida":
                despesa += valor
        
        # Calcula o saldo
        self.saldo = self.salario + receita - despesa
        self.saldoEntradas = receita
        self.saldoSaidas = despesa




        
    def homepage(self):
        self.obter_saldo()
        self.janela.geometry("1024x720")
        self.janela.configure(fg_color=fundo)
        #frame esquerda------------------
        self.frameEsquerda = CTkFrame(master=self.janela,width=223,height=666,fg_color=fundo_cima,corner_radius=30)
        self.frameEsquerda.place(x=20,y=25)
        #Imagem Usuario
        self.imguser = CTkImage(light_image=Image.open("assets\Frame esquerda\ImgFrameEsquerda.png"),dark_image=Image.open("assets\Frame esquerda\ImgFrameEsquerda.png"),size=(192,96))
        self.imguserLabel = CTkLabel(master=self.frameEsquerda,image=self.imguser,text=None)
        self.imguserLabel.place(x=15,y=15)
        
        #textos Dinamicos Img user
        
        CTkLabel(master=self.imguserLabel,text=f"{self.username} ".upper(),text_color="black",font=("Impact",20),fg_color=fundo_cima_utils,bg_color="transparent").place(x=87,y=20)
        CTkLabel(master=self.imguserLabel,text=f"R${self.salario}  ".upper(),text_color="white",font=("Impact",15),fg_color=fundo_cima_utils,bg_color="transparent").place(x=90,y=55)
        
        
        #Botões esquerda
        
        self.btnhome = CTkButton(master=self.frameEsquerda,height=70,width=150,corner_radius=30,text="Inicial".upper(),font=("Impact",20),command=self.homepage)
        self.btnhome.place(x=35,y=135)
        self.btncadastro = CTkButton(master=self.frameEsquerda,height=70,width=150,corner_radius=30,text="cadastro".upper(),font=("Impact",20),command=self.cadastropage)
        self.btncadastro.place(x=35,y=215)
        self.btnentradas = CTkButton(master=self.frameEsquerda,height=70,width=150,corner_radius=30,text="entradas".upper(),font=("Impact",20),command=self.entradaspage)
        self.btnentradas.place(x=35,y=295) 
        self.btnsaidas = CTkButton(master=self.frameEsquerda,height=70,width=150,corner_radius=30,text="saidas".upper(),font=("Impact",20),command=self.saidaspage)
        self.btnsaidas.place(x=35,y=375) 
        self.btntodas = CTkButton(master=self.frameEsquerda,height=70,width=150,corner_radius=30,text="todos".upper(),font=("Impact",20),command=self.todospage)
        self.btntodas.place(x=35,y=455) 
        self.btneditar = CTkButton(master=self.frameEsquerda,height=70,width=150,corner_radius=30,text="dados".upper(),font=("Impact",20),command=self.dadospage)
        self.btneditar.place(x=35,y=535) 
        
        
        
        #frame central---------------------
        self.posxframecentral = 280
        
        self.framecentro = CTkFrame(master=self.janela,width=710,height=460,fg_color=fundo_cima,corner_radius=30)
        self.framecentro.place(x=self.posxframecentral,y=160)
        
        
        CTkLabel(master=self.framecentro,text=self.saldo,text_color="black",font=("Impact",25),fg_color=fundo_cima,bg_color=fundo_cima).place(x=50,y=100)
        CTkLabel(master=self.framecentro,text=self.saldoEntradas,text_color="black",font=("Impact",25),fg_color=fundo_cima,bg_color=fundo_cima).place(x=50,y=150)
        CTkLabel(master=self.framecentro,text=self.saldoSaidas,text_color="black",font=("Impact",25),fg_color=fundo_cima,bg_color=fundo_cima).place(x=50,y=200)
        
        CTkButton(master=self.janela,height=60,corner_radius=30,text="Sair",command=self.janela.destroy).place(x=840,y=640)
        
        
        #decoração frame central
        self.planeta = CTkImage(light_image=Image.open("assets\decor\planeta.png"),dark_image=Image.open("assets\decor\planeta.png"),size=(70,40))
        self.planetalabel = CTkLabel(master=self.janela,image=self.planeta,text=None)
        self.planetalabel.place(x=946,y=27)
    
    
    #enviar cadastro
    def enviarCadastro(self):
        #tratamento de erro 
        if self.tipoEntradaGet.get() == "Tipo de Registro" or len(self.dataDiaGet.get()) >2:
            def fechar_erro():
                erro_toplevel.destroy()
            # Cria a janela pop-up
            erro_toplevel = CTkToplevel()
            erro_toplevel.title("Erro")
            erro_toplevel.geometry("300x150")
            # Torna a janela de erro uma janela temporária
            erro_toplevel.transient(self.janela)
            # Define a janela de erro como sempre no topo
            erro_toplevel.attributes("-topmost", True)
            # Foca na janela de erro
            erro_toplevel.grab_set()

            # Adiciona uma etiqueta com a mensagem de erro
            label = CTkLabel(master=erro_toplevel, text="Data ou Tipo de Registro incorretos", text_color="red", font=("Arial", 14))
            label.pack(pady=20)

            # Adiciona um botão para fechar o pop-up
            botao_fechar = CTkButton(master=erro_toplevel, text="Fechar", command=fechar_erro)
            botao_fechar.pack(pady=10)
            return
        
        
        #coleta dos dados
        self.descricaoRegistro = self.descricaoRegistroGet.get()
        self.valorRegistro = float(self.valorRegistroGet.get().replace(',', '.'))
        self.dataDia = self.dataDiaGet.get()
        self.dataMes = self.dataMesGet.get()
        self.dataAno = self.dataAnoGet.get()
        self.tipoEntrada = self.tipoEntradaGet.get()
        self.categoriaEntrada = self.categoriaEntradaGet.get()
        
        
        # inserindo os dados
        
        planilha = openpyxl.load_workbook(f"anotacoes/{self.username}_controle_financeiro.xlsx")
        planilhaaberta = planilha.active
        planilhaaberta.append([self.dataDia,self.dataMes,self.dataAno,self.descricaoRegistro,self.categoriaEntrada,self.valorRegistro,self.tipoEntrada])
        planilha.save(f"anotacoes/{self.username}_controle_financeiro.xlsx")
        
        self.cadastropage()
        
        
    
    def cadastropage(self):
        print("cadastro")
        #frame central cadastro
        self.framecentro = CTkFrame(master=self.janela,width=710,height=460,fg_color=fundo_cima,corner_radius=30)
        self.framecentro.place(x=self.posxframecentral,y=160)
        #CTkLabel(master=self.framecentro,text="Testando muito",text_color="black").place(x=15,y=150)
        #Descrição Registro
        self.descricaoRegistroGet = CTkEntry(master=self.framecentro,placeholder_text="Descrição do Registro",width=400)
        self.descricaoRegistroGet.place(x=70,y=100)
        #Valor Registro
        self.valorRegistroGet = CTkEntry(master=self.framecentro,width=100,placeholder_text="Valor")
        self.valorRegistroGet.place(x=500,y=100)
        #Data Dia
        self.dataDiaGet = CTkEntry(master=self.framecentro,width=50,placeholder_text="Dia")
        self.dataDiaGet.place(x=70,y=150)
        #Data Mês
        self.dataMesGet = CTkOptionMenu(master=self.framecentro,values=["Janeiro","Fevereiro","Março","Abril","Junho","Julho","Agosto","Setembro","Outubro","Novembro","Dezembro"])
        self.dataMesGet.place(x=135,y=150)
        self.dataMesGet.set("Mês")
        #Data Ano
        self.dataAnoGet = CTkOptionMenu(master=self.framecentro,values=["2020","2021","2022","2023","2024","2025","2026","2027"])
        self.dataAnoGet.place(x=300,y=150)
        self.dataAnoGet.set("Ano")
        #Tipo entrada
        self.tipoEntradaGet = CTkOptionMenu(master=self.framecentro,values=["Entrada","Saida"])
        self.tipoEntradaGet.place(x=500,y=150)
        self.tipoEntradaGet.set("Tipo de Registro")
        #Categoria entrada
        self.categoriaEntradaGet = CTkOptionMenu(master=self.framecentro,values=["Lazer","Contas","Saúde","Comidas","Deposito","Salario"])
        self.categoriaEntradaGet.place(x=300,y=220)   
        self.categoriaEntradaGet.set("Categoria") 
        CTkButton(master=self.framecentro,command=self.enviarCadastro).place(x=500,y=300)
            
        pass
    def entradaspage(self):
        pass
    def saidaspage(self):
        pass
    def todospage(self):
        pass
    def dadospage(self):
        print("dados")
        #frame central dados
        self.framecentro = CTkFrame(master=self.janela,width=710,height=460,fg_color=fundo_cima,corner_radius=30)
        self.framecentro.place(x=self.posxframecentral,y=160)
        #nsalario
        CTkLabel(master=self.framecentro,text="Salario aumentou ?",font=("Impact",40),text_color="black").place(x=30,y=30)
        self.novoSalario = CTkEntry(master=self.framecentro,placeholder_text="Digite aqui o seu novo salario !",width=300)
        self.novoSalario.place(x=40,y=90)
        
        #nsenha
        CTkLabel(master=self.framecentro,text="Senha antiga ?",font=("Impact",40),text_color="black").place(x=30,y=130)
        self.novaSenha = CTkEntry(master=self.framecentro,placeholder_text="Digite aqui sua nova senha",width=300)
        self.novaSenha.place(x=40,y=190)
        
        #nuser
        CTkLabel(master=self.framecentro,text="Cansou do mesmo usuario ?",font=("Impact",40),text_color="black").place(x=30,y=230)
        self.novoUsuario = CTkEntry(master=self.framecentro,placeholder_text="Digite aqui como prefere ser chamado",width=300)
        self.novoUsuario.place(x=40,y=290)
        
        #nplanilha
        CTkLabel(master=self.framecentro,text="Planilha cheia demais ?",font=("Impact",40),text_color="black").place(x=30,y=330)
        CTkButton(master=self.framecentro,text="Aperte aqui e crie uma nova",corner_radius=30,width=160,height=35).place(x=300,y=390)
        
        
        
        
        
        
        
# win = Janela("sei",125000)

# win.run()