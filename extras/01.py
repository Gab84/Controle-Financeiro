import os
import openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo

# Nome do arquivo onde os dados dos usuários serão salvos
FILENAME = "usuarios.txt"

# Função para carregar os usuários do arquivo
def carregar_usuarios():
    usuarios = {}
    if os.path.exists(FILENAME):
        with open(FILENAME, "r") as file:
            for line in file:
                parts = line.strip().split(",")
                if len(parts) == 3:
                    username, password, salario = parts
                    usuarios[username] = (password, salario)
    return usuarios

# Função para salvar os usuários no arquivo
def salvar_usuarios(usuarios):
    with open(FILENAME, "w") as file:
        for username, (password, salario) in usuarios.items():
            file.write(f"{username},{password},{salario}\n")

# Função para criar uma nova planilha para o usuário com layout de controle financeiro
def criar_planilha_usuario(username, salario):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Controle Financeiro"
    
    # Cabeçalhos padrão para controle financeiro
    headers = ["Dia", "Mês", "Ano", "Descrição", "Categoria", "Valor", "Tipo (Receita/Despesa)"]
    ws.append(headers)
    
    # Adiciona uma tabela com filtros
    tab = Table(displayName="ControleFinanceiro", ref="A1:G100000")
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    tab.tableStyleInfo = style
    ws.add_table(tab)
    
    # Adiciona cédulas para o salário e o saldo
    ws["I1"] = "Salário"
    ws["I2"] = salario
    
    ws["J1"] = "Saldo"
    # Fórmula para calcular o saldo conforme "Entrada" e "Saída"
    ws["J2"] = "=I2 + SUMIF(G2:G100000, \"Entrada\", F2:F100000) - SUMIF(G2:G100000, \"Saida\", F2:F100000)"
    
    # Salva o arquivo
    wb.save(f"anotacoes/{username}_controle_financeiro.xlsx")

# Função para cadastrar um novo usuário
def cadastrar_usuario(usuarios):
    username = input("Digite o nome de usuário: ")
    if username in usuarios:
        print("Usuário já existe!")
    else:
        password = input("Digite a senha: ")
        salario = input("Digite o salário mensal: ")
        usuarios[username] = (password, salario)
        salvar_usuarios(usuarios)
        criar_planilha_usuario(username, salario)
        print("Usuário cadastrado com sucesso!")

# Função para realizar o login de um usuário
def login_usuario(usuarios):
    username = input("Digite o nome de usuário: ")
    if username not in usuarios:
        print("Usuário não encontrado!")
    else:
        password = input("Digite a senha: ")
        if usuarios[username][0] == password:
            print("Login bem-sucedido!")
            menu_anotacoes(username)
        else:
            print("Senha incorreta!")

# Função para visualizar as anotações de um usuário
def visualizar_anotacoes(username):
    try:
        wb = openpyxl.load_workbook(f"anotacoes/{username}_controle_financeiro.xlsx")
        ws = wb.active
        print("\nSeu controle financeiro:")
        if ws.max_row == 1:
            print("Nenhuma anotação encontrada.")
        else:
            for row in ws.iter_rows(min_row=2, values_only=True):
                print(row)
    except FileNotFoundError:
        print("Arquivo de controle financeiro não encontrado!")

# Função para adicionar uma nova anotação
def adicionar_anotacao(username):
    dia, mes, ano = map(int, input("Digite a data (dd/mm/aaaa): ").split('/'))
    descricao = input("Digite a descrição: ")
    categoria = input("Digite a categoria: ")
    valor = float(input("Digite o valor: "))
    tipo = input("Digite o tipo (Receita/Despesa): ")
    
    wb = openpyxl.load_workbook(f"anotacoes/{username}_controle_financeiro.xlsx")
    ws = wb.active
    ws.append([dia, mes, ano, descricao, categoria, valor, tipo])
    wb.save(f"anotacoes/{username}_controle_financeiro.xlsx")
    print("Anotação adicionada com sucesso!")
    


# Função para atualizar informações do usuário
def atualizar_informacoes(username, usuarios):
    while True:
        print("\nOpções de Atualização")
        print("1. Alterar Nome de Usuário")
        print("2. Alterar Senha")
        print("3. Alterar Salário")
        print("4. Voltar")
        escolha = input("Escolha uma opção: ")

        if escolha == "1":
            novo_username = input("Digite o novo nome de usuário: ")
            if novo_username in usuarios:
                print("Nome de usuário já existe!")
            else:
                usuarios[novo_username] = usuarios.pop(username)
                os.rename(f"anotacoes/{username}_controle_financeiro.xlsx", f"anotacoes/{novo_username}_controle_financeiro.xlsx")
                salvar_usuarios(usuarios)
                print("Nome de usuário alterado com sucesso!")
                username = novo_username

        elif escolha == "2":
            nova_senha = input("Digite a nova senha: ")
            usuarios[username] = (nova_senha, usuarios[username][1])
            salvar_usuarios(usuarios)
            print("Senha alterada com sucesso!")

        elif escolha == "3":
            novo_salario = input("Digite o novo salário: ")
            usuarios[username] = (usuarios[username][0], novo_salario)
            # Atualiza a planilha com o novo salário
            wb = openpyxl.load_workbook(f"anotacoes/{username}_controle_financeiro.xlsx")
            ws = wb.active
            ws["I2"] = novo_salario
            wb.save(f"anotacoes/{username}_controle_financeiro.xlsx")
            salvar_usuarios(usuarios)
            print("Salário alterado com sucesso!")

        elif escolha == "4":
            break

        else:
            print("Opção inválida! Tente novamente.")

# Menu para as anotações do usuário
def menu_anotacoes(username):
    while True:
        print("\nMenu de Controle Financeiro")
        print("1. Visualizar Controle Financeiro")
        print("2. Adicionar Registro")
        print("3. Sair")
        escolha = input("Escolha uma opção: ")

        if escolha == "1":
            visualizar_anotacoes(username)
        elif escolha == "2":
            adicionar_anotacao(username)
        elif escolha == "3":
            break
        else:
            print("Opção inválida! Tente novamente.")

# Função principal
def main():
    if not os.path.exists("anotacoes"):
        os.makedirs("anotacoes")

    usuarios = carregar_usuarios()
    print(usuarios)
    
    while True:
        print("\nSistema de Login")
        print("1. Cadastrar")
        print("2. Login")
        print("3. Sair")
        escolha = input("Escolha uma opção: ")

        if escolha == "1":
            cadastrar_usuario(usuarios)
        elif escolha == "2":
            login_usuario(usuarios)
        elif escolha == "3":
            break
        else:
            print("Opção inválida! Tente novamente.")

if __name__ == "__main__":
    main()
