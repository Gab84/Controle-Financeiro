from customtkinter import *
from PIL import Image, ImageTk

import openpyxl
from tkinter import ttk
import pandas as pd

from matplotlib.figure import Figure
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg


import platform

fontehome = ("Impact",25)
fontesalario = ("Impact",15)


def check_os():
    os_name = platform.system()
    if os_name == "Windows":
        fontehome = ("Impact",25)
        fontesalario = ("Impact",15)
    elif os_name == "Linux":
        fontehome = ("Arial",25)
        fontesalario = ("Arial",15)
    else:
        print(f"Estou usando {os_name}")


#cores
fundo = "#4D3D14"
fundo_cima = "#F7EED7"
botoes = "#F4C68F"
fundo_cima_utils= "#A28F61"
#auxiliar





class Janela:
    def __init__(self,username,salario) -> None:
        self.janela = CTk()
        self.username = username
        self.salario = salario
        check_os()
        self.homepage()
        self.barradeuso()
        
    def run(self):
        self.janela.mainloop()
    
    def obter_saldo(self):
        
        # Carrega a planilha existente
        wb = openpyxl.load_workbook(f"./planilha_anotacoes/{self.username}_controle_financeiro.xlsx", data_only=True)
        ws = wb.active
        
        # Lê o valor do salário
        self.salario = ws["I2"].value
        
        # Inicializa as variáveis de receita e despesa
        receita = 0
        despesa = 0
        
        # Itera sobre as linhas da planilha e acumula as receitas e despesas
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=6, max_col=7, values_only=True):
            valor, tipo = row
            if tipo == "Entrada":
                receita += valor
            elif tipo == "Saida":
                despesa += valor
        
        # Calcula o saldo
        self.saldo = self.salario + receita - despesa
        self.saldoEntradas = receita
        self.saldoSaidas = despesa
    
    

    
    
    def get_column_widths(self):
        column_widths = {}
        for i, column in enumerate(self.df.columns):
            if i == 6:  # Verifica se é a coluna 7 (índice 6)
                column_widths[column] = 100  # Define uma largura maior
            if i == 5:  # Verifica se é a coluna 7 (índice 6)
                column_widths[column] = 50  # Define uma largura maior
                
            else:
                column_widths[column] = 50
        return column_widths
    
    def apply_filter(self, filter_type):
        self.filter_var.set(filter_type)
        self.update_table()

    def initialize_table(self):
        self.tree["column"] = list(self.df.columns)
        
        for column in self.tree["column"]:
            self.tree.heading(column, text=column)
            self.tree.column(column, minwidth=50, width=self.column_widths[column], stretch=True)
        
        self.update_table()

        
    def apply_sort(self, sort_type):
        self.sort_order[sort_type] = not self.sort_order[sort_type]
        self.sort_by = sort_type
        self.update_table()   
        
    def on_treeview_click(self, event):
        region = self.tree.identify("region", event.x, event.y)
        if region == "heading":
            column = self.tree.identify_column(event.x)
            column_name = self.tree.column(column, option="id")
            if column_name in self.sort_order:
                self.apply_sort(column_name)
                  

    def load_excel(self):
        # Carregar o Excel e inverter a ordem das linhas
        self.df = pd.read_excel(f"./planilha_anotacoes/{self.username}_controle_financeiro.xlsx").fillna("")
        self.df = self.df.iloc[::-1]  # Inverte a ordem das linhas
        self.df = self.df.iloc[:, :7]  # Seleciona apenas as colunas A a G (primeiras 7 colunas)


    def update_table(self):
        self.load_excel()
        self.tree.delete(*self.tree.get_children())
        
        filtered_df = self.df
        filter_type = self.filter_var.get()
        
        if "Tipo (Receita/Despesa)" in self.df.columns:
            if filter_type == "Entradas":
                filtered_df = self.df[self.df["Tipo (Receita/Despesa)"] == "Entrada"]
            elif filter_type == "Saidas":
                filtered_df = self.df[self.df["Tipo (Receita/Despesa)"] == "Saida"]
        else:
            print("A coluna 'Tipo (Receita/Despesa)' não foi encontrada no DataFrame.")
        
        # Converta as colunas "Dia", "Mês" e "Ano" para string antes de ordenar
        for col in ["Dia", "Mês", "Ano"]:
            if col in filtered_df.columns:
                filtered_df[col] = filtered_df[col].astype(str)

        if hasattr(self, 'sort_by'):
            if self.sort_by == "Dia" and "Dia" in filtered_df.columns:
                filtered_df = filtered_df.sort_values(by="Dia", ascending=self.sort_order["Dia"])
            elif self.sort_by == "Mês" and "Mês" in filtered_df.columns:
                filtered_df = filtered_df.sort_values(by="Mês", ascending=self.sort_order["Mês"])
            elif self.sort_by == "Ano" and "Ano" in filtered_df.columns:
                filtered_df = filtered_df.sort_values(by="Ano", ascending=self.sort_order["Ano"])

        for row in filtered_df.to_numpy().tolist():
            self.tree.insert("", "end", values=row)

        
    def homepage(self):
        #Janela ---------
        self.obter_saldo()
        self.janela.geometry("1024x720")
        self.janela.configure(fg_color=fundo)
        
        
        #frame central---------------------
        self.posxframecentral = 280
        
        self.framecentro = CTkFrame(master=self.janela,width=710,height=460,fg_color=fundo_cima,corner_radius=30)
        self.framecentro.place(x=self.posxframecentral,y=160)
        

        # Configuração do estilo do ttk para combinar com CustomTkinter
        style = ttk.Style()
        style.theme_use("clam")
        style.configure("Custom.Treeview",
                        background="#D3D3D3",
                        foreground="black",
                        rowheight=25,
                        fieldbackground="#D3D3D3")
        
        style.configure("Custom.Treeview.Heading",
                background="#4D4D4D",
                foreground="white",
                font=("Helvetica", 10, "bold"))
        
        style.map('Custom.Treeview',
                  background=[('selected', '#347083')],
                  foreground=[('selected', 'white')])

        # Criação do Treeview
        self.tree = ttk.Treeview(self.framecentro, show="headings", height=5)
        self.tree.place(x=30,y=280)
        
        self.tree.bind("<Button-1>", self.on_treeview_click)
        
        self.filter_var = StringVar(value="Todos")
        self.sort_order = {"Dia": False, "Mês": False, "Ano": False}
        
        self.load_excel()
        self.column_widths = self.get_column_widths()
        self.initialize_table()
        
        
        self.all_button = CTkButton(self.framecentro, text="Todos", command=lambda: self.apply_filter("Todos"),width=40)
        self.all_button.place(x=30,y=420)
        
        self.entries_button = CTkButton(self.framecentro, text="Entradas", command=lambda: self.apply_filter("Entradas"),width=40)
        self.entries_button.place(x=90,y=420)
        
        self.exits_button = CTkButton(self.framecentro, text="Saídas", command=lambda: self.apply_filter("Saidas"),width=40)
        self.exits_button.place(x=170,y=420)
        
        
        
        
        #frame esquerda------------------
        self.frameEsquerda = CTkFrame(master=self.janela,width=223,height=666,fg_color=fundo_cima,corner_radius=30)
        self.frameEsquerda.place(x=20,y=25)
        #Imagem Usuario
        self.imguser = CTkImage(light_image=Image.open("./imgassets/Frame esquerda/ImgFrameEsquerda.png"),dark_image=Image.open("./imgassets/Frame esquerda/ImgFrameEsquerda.png"),size=(192,96))
        self.imguserLabel = CTkLabel(master=self.frameEsquerda,image=self.imguser,text=None)
        self.imguserLabel.place(x=15,y=15)
        
        #textos Dinamicos Img user
        
        CTkLabel(master=self.imguserLabel,text=f"{self.username} ".upper(),text_color="black",font=fontehome,fg_color=fundo_cima_utils,bg_color="transparent").place(x=87,y=20)
        CTkLabel(master=self.imguserLabel,text=f"R${self.salario}  ".upper(),text_color="white",font=fontesalario,fg_color=fundo_cima_utils,bg_color="transparent").place(x=90,y=55)
        
        
        #Botões esquerda
        
        self.btnhome = CTkButton(master=self.frameEsquerda,height=70,width=150,corner_radius=30,text="Inicial".upper(),font=fontehome,command=self.homepage)
        self.btnhome.place(x=35,y=135)
        self.btncadastro = CTkButton(master=self.frameEsquerda,height=70,width=150,corner_radius=30,text="cadastro".upper(),font=fontehome,command=self.cadastropage)
        self.btncadastro.place(x=35,y=215)
        self.btnentradas = CTkButton(master=self.frameEsquerda,height=70,width=150,corner_radius=30,text="entradas".upper(),font=fontehome,command=self.entradaspage)
        self.btnentradas.place(x=35,y=295) 
        self.btnsaidas = CTkButton(master=self.frameEsquerda,height=70,width=150,corner_radius=30,text="saidas".upper(),font=fontehome,command=self.saidaspage)
        self.btnsaidas.place(x=35,y=375) 
        self.btntodas = CTkButton(master=self.frameEsquerda,height=70,width=150,corner_radius=30,text="todos".upper(),font=fontehome,command=self.todospage)
        self.btntodas.place(x=35,y=455) 
        self.btneditar = CTkButton(master=self.frameEsquerda,height=70,width=150,corner_radius=30,text="dados".upper(),font=fontehome,command=self.dadospage)
        self.btneditar.place(x=35,y=535) 
        
        
        

        
        #mudando isso aqui tbm ............   mudar
        
        CTkLabel(master=self.janela,text=f"Seu saldo atual: R${self.saldo} ",text_color="white",font=fontehome,fg_color=fundo,bg_color=fundo).place(x=280,y=80)
        CTkLabel(master=self.janela,text=f"Dinhiro adicionado: R${self.saldoEntradas} ",text_color="green",font=fontehome,fg_color=fundo,bg_color=fundo).place(x=670,y=80)
        CTkLabel(master=self.janela,text=f"Dinheiro Gasto: R${self.saldoSaidas} ",text_color="red",font=fontehome,fg_color=fundo,bg_color=fundo).place(x=670,y=120)
        CTkLabel(master=self.janela,text=f"Barra de uso ",text_color="white",font=("Arial",20,"italic"),fg_color=fundo,bg_color=fundo).place(x=490,y=120)
        
        
        
        
        
        
        ###
        
        CTkButton(master=self.janela,height=60,corner_radius=30,text="Sair",command=self.janela.destroy).place(x=840,y=640)
        
        

        
        
        #decoração frame central
        self.planeta = CTkImage(light_image=Image.open("./imgassets/decor/planeta.png"),dark_image=Image.open("./imgassets/decor/planeta.png"),size=(70,40))
        self.planetalabel = CTkLabel(master=self.janela,image=self.planeta,text=None)
        self.planetalabel.place(x=946,y=27)
        self.barradeuso()
        
        
        
        
        #parte dos graficos passa para o main  ------ mudar
        
      

        entries_sum = self.df[self.df["Tipo (Receita/Despesa)"] == "Entrada"]["Valor"].sum()
        exits_sum = self.df[self.df["Tipo (Receita/Despesa)"] == "Saida"]["Valor"].sum()
        balance = entries_sum - exits_sum + self.salario
        
        # Garantir que todos os valores sejam não negativos
        entries_sum = max(entries_sum, 0)
        exits_sum = max(exits_sum, 0)
        balance = max(balance, 0)
        
        data = {"Entradas": entries_sum, "Saidas": exits_sum, "Saldo": balance}
        names = list(data.keys())
        values = list(data.values())
        
        fig = Figure(figsize=(5,2))  # Tamanho do gráfico de barras
        fig.patch.set_facecolor(fundo_cima)
        ax = fig.add_subplot(111)
        ax.bar(names, values, color=["#4CAF50", "#F44336", "#FFC107"])  # Cores dos bastões
        ax.set_title("Entradas, Saídas e Saldo")  # Título do gráfico
        
        
        
        canvas = FigureCanvasTkAgg(fig, master=self.framecentro)
        canvas.draw()
        canvas.get_tk_widget().place(x=32,y=20)
        
        
        
        # Categorias a serem mostradas no gráfico de pizza
        categories = ["Lazer", "Contas", "Saúde", "Comidas","Outros"]
        
        # Verifica se a coluna 'Categoria' e 'Valor' existem no DataFrame
        if "Categoria" not in self.df.columns or "Valor" not in self.df.columns:
            print("As colunas 'Categoria' e/ou 'Valor' não foram encontradas no DataFrame.")
            return
        
        # Calcula a soma dos valores para cada categoria
        data = self.df[self.df["Categoria"].isin(categories)].groupby("Categoria")["Valor"].sum()
        
        # Garantir que todos os valores sejam não negativos
        data = data.clip(lower=0)
        
        labels = data.index.tolist()
        values = data.tolist()
        colors = ["#4CAF50", "#FFC107", "#FF5722", "#03A9F4", "#E91E63"]  # Cores das fatias

        fig = Figure(figsize=(2.5,2))  # Tamanho do gráfico de pizza
        fig.patch.set_facecolor(fundo_cima)
        ax = fig.add_subplot(111)
        ax.pie(values, labels=labels, colors=colors, autopct='%1.1f%%', startangle=90)  # Configuração do gráfico de pizza
        ax.set_title("Distribuição Financeira")  # Título do gráfico
        
        canvas = FigureCanvasTkAgg(fig, master=self.framecentro)
        canvas.draw()
        canvas.get_tk_widget().place(x=430,y=250)
        
        
        
        
        
        
        
        
        
        # from matplotlib.figure import Figure
        # from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
        
        
        #fim dos graficos 
        
        
        
        
    
    def barradeuso(self):
        
        self.barra = CTkProgressBar(master=self.janela,height=30,corner_radius=0)
        self.barra.place(x=280,y=120)
        
        saldo = self.salario + self.saldoEntradas
        saidas = (self.saldoSaidas*100)
        porcentagem = (saidas/saldo)/100
        self.barra.set(porcentagem)
    
    #enviar cadastro
    def enviarCadastro(self):
        
        # tratamento de erro 
        if self.tipoEntradaGet.get() == "Tipo de Registro" or len(self.dataDiaGet.get()) > 2 or len(self.dataDiaGet.get()) < 1 or self.dataMesGet.get() == "Mês" or self.dataAnoGet.get() == "Ano":
            def fechar_erro():
                erro_toplevel.destroy()
            # Cria a janela pop-up
            erro_toplevel = CTkToplevel()
            erro_toplevel.title("Erro")
            erro_toplevel.geometry("300x150")
            # Torna a janela de erro uma janela temporária
            erro_toplevel.transient(self.janela)
            # Define a janela de erro como sempre no topo
            erro_toplevel.attributes("-topmost", True)
            # Foca na janela de erro
            erro_toplevel.grab_set()

            # Adiciona uma etiqueta com a mensagem de erro
            label = CTkLabel(master=erro_toplevel, text="Preencha todos os campos corretamente.", text_color="red", font=("Arial", 14))
            label.pack(pady=20)

            # Adiciona um botão para fechar o pop-up
            botao_fechar = CTkButton(master=erro_toplevel, text="Fechar", command=fechar_erro)
            botao_fechar.pack(pady=10)
            return
        
        # coleta dos dados
        self.descricaoRegistro = self.descricaoRegistroGet.get()
        self.valorRegistro = float(self.valorRegistroGet.get().replace(',', '.'))
        self.dataDia = int(self.dataDiaGet.get())
        self.dataMes = self.dataMesGet.get()
        self.dataAno = int(self.dataAnoGet.get())
        self.tipoEntrada = self.tipoEntradaGet.get()
        self.categoriaEntrada = self.categoriaEntradaGet.get()
        
         # mudar
        if self.categoriaEntrada == "Categoria":
            self.categoriaEntrada = "Outros"
        
        
        ##
        
        self.barradeuso()
        
        
        # inserindo os dados
        
        planilha = openpyxl.load_workbook(f"./planilha_anotacoes/{self.username}_controle_financeiro.xlsx")
        planilhaaberta = planilha.active
        planilhaaberta.append([self.dataDia,self.dataMes,self.dataAno,self.descricaoRegistro,self.categoriaEntrada,self.valorRegistro,self.tipoEntrada])
        planilha.save(f"./planilha_anotacoes/{self.username}_controle_financeiro.xlsx")
        
        self.cadastropage()
        
        
        
    
    def cadastropage(self):
        
        print("cadastro")
        #frame central cadastro
        self.framecentro = CTkFrame(master=self.janela,width=710,height=460,fg_color=fundo_cima,corner_radius=30)
        self.framecentro.place(x=self.posxframecentral,y=160)
        #CTkLabel(master=self.framecentro,text="Testando muito",text_color="black").place(x=15,y=150)
        #Descrição Registro
        self.descricaoRegistroGet = CTkEntry(master=self.framecentro,placeholder_text="Descrição do Registro",width=400)
        self.descricaoRegistroGet.place(x=70,y=100)
        #Valor Registro
        self.valorRegistroGet = CTkEntry(master=self.framecentro,width=100,placeholder_text="Valor")
        self.valorRegistroGet.place(x=500,y=100)
        #Data Dia
        self.dataDiaGet = CTkEntry(master=self.framecentro,width=50,placeholder_text="Dia")
        self.dataDiaGet.place(x=70,y=150)
        #Data Mês
        self.dataMesGet = CTkOptionMenu(master=self.framecentro,values=["Janeiro","Fevereiro","Março","Abril","Junho","Julho","Agosto","Setembro","Outubro","Novembro","Dezembro"])
        self.dataMesGet.place(x=135,y=150)
        self.dataMesGet.set("Mês")
        #Data Ano
        self.dataAnoGet = CTkOptionMenu(master=self.framecentro,values=["2020","2021","2022","2023","2024","2025","2026","2027"])
        self.dataAnoGet.place(x=300,y=150)
        self.dataAnoGet.set("Ano")
        #Tipo entrada
        self.tipoEntradaGet = CTkOptionMenu(master=self.framecentro,values=["Entrada","Saida"])
        self.tipoEntradaGet.place(x=500,y=150)
        self.tipoEntradaGet.set("Tipo de Registro")
        #Categoria entrada
        self.categoriaEntradaGet = CTkOptionMenu(master=self.framecentro,values=["Lazer","Contas","Saúde","Comidas","Deposito","Salario"])
        self.categoriaEntradaGet.place(x=300,y=220)   
        self.categoriaEntradaGet.set("Categoria") 
        CTkButton(master=self.framecentro,command=self.enviarCadastro).place(x=500,y=300)
        
        pass
    def entradaspage(self):
        pass
    def saidaspage(self):
        pass
    def todospage(self):
        pass
    def dadospage(self):
        print("dados")
        #frame central dados
        self.framecentro = CTkFrame(master=self.janela,width=710,height=460,fg_color=fundo_cima,corner_radius=30)
        self.framecentro.place(x=self.posxframecentral,y=160)
        #nsalario
        CTkLabel(master=self.framecentro,text="Salario aumentou ?",font=("Impact",40),text_color="black").place(x=30,y=30)
        self.novoSalario = CTkEntry(master=self.framecentro,placeholder_text="Digite aqui o seu novo salario !",width=300)
        self.novoSalario.place(x=40,y=90)
        
        #nsenha
        CTkLabel(master=self.framecentro,text="Senha antiga ?",font=("Impact",40),text_color="black").place(x=30,y=130)
        self.novaSenha = CTkEntry(master=self.framecentro,placeholder_text="Digite aqui sua nova senha",width=300)
        self.novaSenha.place(x=40,y=190)
        
        #nuser
        CTkLabel(master=self.framecentro,text="Cansou do mesmo usuario ?",font=("Impact",40),text_color="black").place(x=30,y=230)
        self.novoUsuario = CTkEntry(master=self.framecentro,placeholder_text="Digite aqui como prefere ser chamado",width=300)
        self.novoUsuario.place(x=40,y=290)
        
        #nplanilha
        CTkLabel(master=self.framecentro,text="Planilha cheia demais ?",font=("Impact",40),text_color="black").place(x=30,y=330)
        CTkButton(master=self.framecentro,text="Aperte aqui e crie uma nova",corner_radius=30,width=160,height=35).place(x=300,y=390)
        
        
        
        
        
        
"""      
win = Janela("gabs",125000)

win.run()"""