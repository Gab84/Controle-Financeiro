from customtkinter import *
import openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo
from PIL import Image, ImageTk

import os

from Main import Janela as jnl

FILENAME = "usuarios.txt"

class ImageHoverButton(CTkLabel):
    def __init__(self, master, image_path, hover_image_path, command=None, **kwargs):
        super().__init__(master, text="", **kwargs)  # Define o texto como vazio
        self.command = command

        # Carregar imagens
        default_image = Image.open(image_path)
        hover_image = Image.open(hover_image_path)

        # Criar CTkImage com o tamanho original
        self.default_image = CTkImage(light_image=default_image, size=default_image.size)
        self.hover_image = CTkImage(light_image=hover_image, size=hover_image.size)

        # Configurar o label com a imagem padrão
        self.configure(image=self.default_image)
        self.bind("<Enter>", self.on_enter)
        self.bind("<Leave>", self.on_leave)
        self.bind("<Button-1>", self.on_click)

    def on_enter(self, event):
        self.configure(image=self.hover_image)

    def on_leave(self, event):
        self.configure(image=self.default_image)

    def on_click(self, event):
        if self.command:
            self.command()


class Janela:
    def __init__(self):
        self.janela = CTk()
        #self.jlogin = CTk()
        #self.jcadastro = CTk()
        carregar_user = self.carregar_usuarios()
        print(self.usuarios)
        
    def carregar_usuarios(self):
        self.usuarios = {}
        if os.path.exists(FILENAME):
            with open(FILENAME, "r") as file:
                for line in file:
                    parts = line.strip().split(",")
                    if len(parts) == 3:
                        self.usuario, self.senha,self.salario = parts
                        self.usuarios[self.usuario] = (self.senha,self.salario)
                    
    def salvar_usuarios(self):
        with open(FILENAME, "w") as file:
            for self.nusuariox, (self.password,self.salario) in self.usuarios.items():
                file.write(f"{self.nusuariox},{self.password},{self.salario}\n")

    def criar_planilha_usuario(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Controle Financeiro"
        
        # Cabeçalhos padrão para controle financeiro
        headers = ["Dia", "Mês", "Ano", "Descrição", "Categoria", "Valor", "Tipo (Receita/Despesa)"]
        ws.append(headers)
        
        # Adiciona uma tabela com filtros
        tab = Table(displayName="ControleFinanceiro", ref="A1:G100000")
        style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=True)
        tab.tableStyleInfo = style
        ws.add_table(tab)
        
        # Adiciona cédulas para o salário e o saldo
        ws["I1"] = "Salário"
        ws["I2"] = float(self.salario)
        
        ws["J1"] = "Saldo"
        # Fórmula para calcular o saldo conforme "Entrada" e "Saída"
        ws["J2"] = "=I2 + SUMIF(G2:G100000, \"Entrada\", F2:F100000) - SUMIF(G2:G100000, \"Saida\", F2:F100000)"
        
        # Salva o arquivo
        wb.save(f"anotacoes/{self.nusuariox}_controle_financeiro.xlsx")

    def cadastrar_usuario(self):
        self.nusuariox = self.nusuario.get()
        if self.nusuariox in self.usuarios:
            print("Usuário já existe!")
        else:
            self.password = self.nsenha.get()
            self.salario = self.nsalario.get()
            self.usuarios[self.nusuariox] = (self.password,self.salario)
            self.salvar_usuarios()
            self.criar_planilha_usuario()
            print("Usuário cadastrado com sucesso!")

    
    def home(self):
        self.janela.geometry("620x400")
        self.janela._set_appearance_mode("dark")
        self.janela.resizable(False,False)
        self.janela.title("Tela Inicial")
        self.janela.iconbitmap("icons\icons8-casa-48.ico")
        self.janela.configure(fg_color="#F2F2F2",bg_color="transparent")
        
        self.frameEscolha = CTkFrame(master=self.janela,width=500,height=320,fg_color="#1A1926",corner_radius=20,)
        self.homeImage = CTkImage(light_image=Image.open("images\homeImage.png"),dark_image=Image.open("images\homeImage.png"),size=(280,300))
        self.lbHomeImage = CTkLabel(master=self.frameEscolha,image=self.homeImage,text=None)
        self.lbHomeImage.place(x=10,y=15)
        
        self.cadastroImage = "Bottoes Img\iregistroImg.png"
        self.cadastroImagehover = "Bottoes Img\hover_img\hover_cadastro.png"

        self.loginImage = "Bottoes Img\loginImg.png"
        self.loginImagehover = "Bottoes Img\hover_img\hover_login.png"
        
        
        self.voltImage = "Bottoes Img\homeImg.png"
        self.voltImagehove = "Bottoes Img\hover_img\hover_home.png"
        
        self.Imgentrar = "Bottoes Img\ImgEntrar.png"
        self.Imgentrarhover = "Bottoes Img\hover_img\ImgEntrarhover.png"

        self.ImgRegistro = "Bottoes Img\Imgregistro.png"
        self.ImgRegistrohover = "Bottoes Img\hover_img\Imgregistrohover.png"
        
        ImageHoverButton(master=self.frameEscolha,
                         image_path=self.loginImage,
                         hover_image_path=self.loginImagehover,
                         command=self.login).place(x=320,y=90)
        
        ImageHoverButton(master=self.frameEscolha,
                         image_path=self.cadastroImage,
                         hover_image_path=self.cadastroImagehover,
                         command=self.cadastro).place(x=320,y=170)
        
        self.animlogin = 700
        self.animcadastro = 700
        self.animEscolha = 60
        
        self.frameLogin = CTkFrame(master=self.janela,width=500,height=320,fg_color="#1A1926",corner_radius=20,)
        self.frameCadastro = CTkFrame(master=self.janela,width=500,height=320,fg_color="#1A1926",corner_radius=20,)
        
        self.frameEscolha.place(x=self.animEscolha,y=40)
        self.frameLogin.place(x=self.animlogin,y=100)
        self.frameCadastro.place(x=self.animcadastro,y=100)
        
        
        
        self.janela.mainloop()
    
    
    def volthome(self):
        
        if self.animlogin < 699:
            self.animlogin +=20
        
        if  self.animcadastro < 699:
            self.animcadastro += 20
        
        if self.animEscolha < 60:
            self.animEscolha += 20.6
            
        if self.animEscolha > 60:
            self.animEscolha = 60
        
        
        if self.animlogin < 699:
            
            self.janela.after(10,self.volthome)
            self.frameLogin.place(x=self.animlogin,y=40)
            self.frameEscolha.place(x=self.animEscolha,y=40)
            self.frameLogin.lift()
            
        if  self.animcadastro < 699:
            
            self.janela.after(10,self.volthome)
            self.frameCadastro.place(x=self.animcadastro,y=40)
            self.frameEscolha.place(x=self.animEscolha,y=40)
            self.frameCadastro.lift()
        
        
        


    def validacao(self):
        
        self.usuariolog = self.usuario.get()
        senha = self.senha.get()
        
        validado1 = False
        validado2 = False
         
        if self.usuariolog in self.usuarios:
            validado1 = True
            
        elif self.usuariolog not in self.usuarios:
            print("Usuario não cadastrado !")
            return
        
        if validado1 and self.usuarios[self.usuariolog][0] == senha:
            validado2 = True
            print("certo")
            self.usuario.delete(0,END)
            self.senha.delete(0,END)
            self.janela.destroy()
            menuuser = jnl(self.usuariolog,self.salario)
            menuuser.run()
             
        else:
            print("usuário ou senha incorretos")
            self.senha.delete(0,END)
        
        
    def login(self):
        
        
        
        self.animlogin -=20
        self.animEscolha -=20

        
        if self.animlogin > 59:
            self.janela.after(10,self.login)
            self.frameLogin.place(x=self.animlogin,y=40)
            self.frameEscolha.place(x=self.animEscolha,y=40)
            self.frameLogin.lift()
            print(self.animlogin)

        if self.animlogin == 660:
             
            self.loginCanva = CTkImage(light_image=Image.open("icons\personagem.png"),dark_image=Image.open("icons\personagem.png"),size=(120,120))
            self.lbLogCanva = CTkLabel(master=self.frameLogin,
                                       image=self.loginCanva,
                                       text=None,
                                       bg_color="#1A1926",fg_color="transparent")
            self.lbLogCanva.place(x=320,y=55)
        
        
        
            
            self.loginCanvalado = CTkImage(light_image=Image.open("icons\imgfundo_login2.png"),dark_image=Image.open("icons\imgfundo_login2.png"),size=(250,320))
            self.lbLogCanvalado = CTkLabel(master=self.frameLogin,
                                       image=self.loginCanvalado,
                                       text=None,
                                       bg_color="#F2F2F2",fg_color="transparent")
            
            self.lbLogCanvalado.place(x=0,y=0)
            

            
            
            ImageHoverButton(master=self.frameLogin,
                             image_path=self.voltImage,
                             hover_image_path=self.voltImagehove,
                             command=self.volthome).place(x=409,y=7)
            
            self.usuario = CTkEntry(master=self.frameLogin,width=200,placeholder_text="Usuario",border_color="#ABA0F2",corner_radius=15,height=20,fg_color="white",placeholder_text_color="black",text_color="black")
            self.usuario.place(x=280,y=185)
            
            self.senha = CTkEntry(master=self.frameLogin,width=200,placeholder_text="Senha",border_color="#ABA0F2",corner_radius=15,height=20,fg_color="white",placeholder_text_color="black",text_color="black")
            self.senha.place(x=280,y=215)
            
            

            ImageHoverButton(master=self.frameLogin,
                             image_path=self.Imgentrar,
                             hover_image_path=self.Imgentrarhover,
                             command=self.validacao,
                             ).place(x=310,y=250)
        
        
        
    def cadastro(self):
        
        
        
        self.animcadastro -=20
        self.animEscolha -=20

        
        if self.animcadastro > 59:
            self.janela.after(10,self.cadastro)
            self.frameCadastro.place(x=self.animcadastro,y=40)
            self.frameEscolha.place(x=self.animEscolha,y=40)
            self.frameCadastro.lift()
            
        if self.animcadastro ==660:
            
            self.loginCanva = CTkImage(light_image=Image.open("icons\personagemcads.png"),dark_image=Image.open("icons\personagemcads.png"),size=(150,100))
            self.lbLogCanva = CTkLabel(master=self.frameCadastro,
                                       image=self.loginCanva,
                                       text=None,
                                       bg_color="#1A1926",fg_color="transparent")
            self.lbLogCanva.place(x=45,y=60)
            
            self.cadCanvalado = CTkImage(light_image=Image.open("icons\imgfundo_cadastro.png"),dark_image=Image.open("icons\imgfundo_cadastro.png"),size=(250,320))
            self.lbcadCanvalado = CTkLabel(master=self.frameCadastro,
                                       image=self.cadCanvalado,
                                       text=None,
                                       bg_color="#F2F2F2",fg_color="transparent",)
            
            self.lbcadCanvalado.place(x=250,y=0)
            
            
            ImageHoverButton(master=self.frameCadastro,
                             image_path=self.voltImage,
                             hover_image_path=self.voltImagehove,
                             command=self.volthome).place(x=7,y=5)
            
            self.nusuario = CTkEntry(master=self.frameCadastro,width=200,placeholder_text="Usuario",border_color="#ABA0F2",corner_radius=15,fg_color="white",height=20,placeholder_text_color="black",text_color="black")
            self.nusuario.place(x=20,y=175)
            
            self.nsenha = CTkEntry(master=self.frameCadastro,width=200,placeholder_text="Senha",border_color="#ABA0F2",corner_radius=15,fg_color="white",height=20,placeholder_text_color="black",text_color="black")
            self.nsenha.place(x=20,y=205)
            
            self.nsalario = CTkEntry(master=self.frameCadastro,width=200,placeholder_text="Salario",border_color="#ABA0F2",corner_radius=15,fg_color="white",height=20,placeholder_text_color="black",text_color="black")
            self.nsalario.place(x=20,y=235)
            
            self.lbLogCanva.lift()
        
            ImageHoverButton(master=self.frameCadastro,
                             image_path=self.ImgRegistro,
                             hover_image_path=self.ImgRegistrohover,
                             command=self.cadastrar_usuario,
                             ).place(x=55,y=270)



wind = Janela()

wind.home()